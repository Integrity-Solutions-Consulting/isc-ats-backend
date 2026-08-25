"""Tests for GET /auth/subscribers and /auth/subscribers/export (marketing-consent
Slice 4 — Marketing role + Excel export).

Covers:
- A user holding ONLY the Marketing role gets 200 on both endpoints.
- The same user gets 403 on /auth/users (outside the Marketing allowlist) — the
  core RBAC guarantee of this slice.
- The count matches exactly the number of users with an active marketing consent
  AND an active candidate profile (seed: active-subscribed, revoked, never-decided,
  active-subscribed-but-soft-deleted-candidate).
- The export has exactly the 4 expected columns, no phone/cedula.
- The export handles a timezone-aware accepted_at without raising (openpyxl gotcha
  regression test).
- An Admin-role user still succeeds on both endpoints.
"""

import io
import uuid
from collections.abc import AsyncGenerator
from datetime import UTC, datetime

import pytest
from httpx import ASGITransport, AsyncClient
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.database import get_session
from app.core.security import create_access_token, hash_password
from app.main import app
from app.modules.auth.application.bootstrap_service import (
    MARKETING_ROLE_NAME,
    assign_role_to_user,
    bootstrap_admin,
)
from app.modules.auth.infrastructure.consents_repository import ConsentsRepository
from app.modules.auth.infrastructure.models import CONSENT_MARKETING, Role, User
from app.modules.auth.infrastructure.repository import UserRepository
from app.modules.org.infrastructure.parameters_repository import ParameterRepository
from app.modules.recruitment.infrastructure.candidate_models import Candidate

POLICY_VERSION = "1.0"
BASE = "/api/v1/auth/subscribers"


@pytest.fixture
async def client(session: AsyncSession) -> AsyncGenerator[AsyncClient, None]:
    async def _override() -> AsyncGenerator[AsyncSession, None]:
        yield session

    app.dependency_overrides[get_session] = _override
    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://test") as ac:
        yield ac
    app.dependency_overrides.clear()


def _bearer(user_id: int, portal: str) -> dict[str, str]:
    token = create_access_token(user_id, extra_claims={"portal": portal})
    return {"Authorization": f"Bearer {token}"}


async def _make_candidate_user(session: AsyncSession, *, tag: str) -> User:
    portal = await ParameterRepository(session).get_by_type_and_code("user_portal", "candidate")
    assert portal is not None
    return await UserRepository(session).add(
        User(
            email=f"sub-{tag}-{uuid.uuid4().hex[:8]}@test.example.com",
            portal_id=portal.id,
            email_verified=True,
        )
    )


async def _make_staff_user(session: AsyncSession) -> User:
    portal = await ParameterRepository(session).get_by_type_and_code("user_portal", "staff")
    assert portal is not None
    return await UserRepository(session).add(
        User(
            email=f"staff-{uuid.uuid4().hex[:8]}@test.example.com",
            password_hash=hash_password("Pass1234!"),
            portal_id=portal.id,
            email_verified=True,
        )
    )


async def _marketing_user(session: AsyncSession) -> User:
    """A staff user holding ONLY the Marketing role (bootstrap creates the role)."""
    await bootstrap_admin(session, f"admin-{uuid.uuid4().hex[:8]}@test.local", "S3cret-pass")
    user = await _make_staff_user(session)
    role = (
        await session.execute(
            select(Role).where(Role.name == MARKETING_ROLE_NAME).where(Role.is_active.is_(True))
        )
    ).scalar_one()
    await assign_role_to_user(session, user.id, role.id)
    await session.flush()
    return user


async def _admin_user(session: AsyncSession) -> User:
    email = f"admin-{uuid.uuid4().hex[:8]}@test.local"
    result = await bootstrap_admin(session, email, "S3cret-pass")
    stmt = select(User).where(User.id == result.user_id)
    return (await session.execute(stmt)).scalar_one()


# ---------------------------------------------------------------------------
# RBAC — Marketing role
# ---------------------------------------------------------------------------


async def test_marketing_role_gets_200_on_count(client: AsyncClient, session: AsyncSession) -> None:
    user = await _marketing_user(session)

    response = await client.get(BASE, headers=_bearer(user.id, "staff"))

    assert response.status_code == 200
    assert response.json() == {"count": 0}


async def test_marketing_role_gets_200_on_export(
    client: AsyncClient, session: AsyncSession
) -> None:
    user = await _marketing_user(session)

    response = await client.get(f"{BASE}/export", headers=_bearer(user.id, "staff"))

    assert response.status_code == 200
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


async def test_marketing_role_gets_403_on_users_route(
    client: AsyncClient, session: AsyncSession
) -> None:
    user = await _marketing_user(session)

    response = await client.get("/api/v1/auth/users", headers=_bearer(user.id, "staff"))

    assert response.status_code == 403


async def test_admin_role_still_succeeds_on_both_endpoints(
    client: AsyncClient, session: AsyncSession
) -> None:
    admin = await _admin_user(session)

    count_response = await client.get(BASE, headers=_bearer(admin.id, "staff"))
    export_response = await client.get(f"{BASE}/export", headers=_bearer(admin.id, "staff"))

    assert count_response.status_code == 200
    assert export_response.status_code == 200


# ---------------------------------------------------------------------------
# Count accuracy
# ---------------------------------------------------------------------------


async def test_count_matches_active_subscribed_candidates_only(
    client: AsyncClient, session: AsyncSession
) -> None:
    marketing_user = await _marketing_user(session)
    repo = ConsentsRepository(session)

    active_user = await _make_candidate_user(session, tag="active")
    revoked_user = await _make_candidate_user(session, tag="revoked")
    undecided_user = await _make_candidate_user(session, tag="undecided")
    soft_deleted_user = await _make_candidate_user(session, tag="soft-deleted")

    session.add(Candidate(user_id=active_user.id, first_name="Ana", last_name="Lopez"))
    session.add(Candidate(user_id=revoked_user.id, first_name="Beto", last_name="Cruz"))
    session.add(Candidate(user_id=undecided_user.id, first_name="Caro", last_name="Diaz"))
    deleted_candidate = Candidate(
        user_id=soft_deleted_user.id, first_name="Dario", last_name="Ega"
    )
    session.add(deleted_candidate)
    await session.flush()

    await repo.grant(
        active_user.id, CONSENT_MARKETING, policy_version=POLICY_VERSION, source="banner"
    )
    await repo.grant(
        revoked_user.id, CONSENT_MARKETING, policy_version=POLICY_VERSION, source="banner"
    )
    await repo.revoke(revoked_user.id, CONSENT_MARKETING, source="banner")
    # undecided_user: never decided — no consent row at all.
    await repo.grant(
        soft_deleted_user.id, CONSENT_MARKETING, policy_version=POLICY_VERSION, source="banner"
    )
    deleted_candidate.is_active = False
    await session.flush()

    response = await client.get(BASE, headers=_bearer(marketing_user.id, "staff"))

    assert response.status_code == 200
    assert response.json() == {"count": 1}


# ---------------------------------------------------------------------------
# Export shape
# ---------------------------------------------------------------------------


async def test_export_has_exactly_the_expected_columns(
    client: AsyncClient, session: AsyncSession
) -> None:
    marketing_user = await _marketing_user(session)
    repo = ConsentsRepository(session)

    subscriber = await _make_candidate_user(session, tag="export")
    session.add(Candidate(user_id=subscriber.id, first_name="Elena", last_name="Fierro"))
    await session.flush()
    await repo.grant(
        subscriber.id, CONSENT_MARKETING, policy_version=POLICY_VERSION, source="banner"
    )

    response = await client.get(f"{BASE}/export", headers=_bearer(marketing_user.id, "staff"))

    assert response.status_code == 200
    workbook = load_workbook(io.BytesIO(response.content))
    sheet = workbook.active
    header = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    assert header == ["Nombres", "Apellidos", "Correo", "Fecha de aceptación"]
    assert "Cédula" not in header
    assert "Teléfono" not in header
    assert "Phone" not in header

    data_row = [cell.value for cell in next(sheet.iter_rows(min_row=2, max_row=2))]
    assert data_row[0] == "Elena"
    assert data_row[1] == "Fierro"
    assert data_row[2] == subscriber.email


async def test_export_handles_timezone_aware_accepted_at_without_raising(
    client: AsyncClient, session: AsyncSession
) -> None:
    """Regression test for the openpyxl tz-aware-datetime crash (design gotcha)."""
    marketing_user = await _marketing_user(session)
    repo = ConsentsRepository(session)

    subscriber = await _make_candidate_user(session, tag="tzaware")
    session.add(Candidate(user_id=subscriber.id, first_name="Gus", last_name="Haro"))
    await session.flush()
    await repo.grant(
        subscriber.id,
        CONSENT_MARKETING,
        policy_version=POLICY_VERSION,
        source="banner",
        accepted_at=datetime.now(UTC),
    )

    response = await client.get(f"{BASE}/export", headers=_bearer(marketing_user.id, "staff"))

    assert response.status_code == 200
    workbook = load_workbook(io.BytesIO(response.content))
    sheet = workbook.active
    data_row = [cell.value for cell in next(sheet.iter_rows(min_row=2, max_row=2))]
    assert data_row[3] is not None
    assert data_row[3].tzinfo is None
